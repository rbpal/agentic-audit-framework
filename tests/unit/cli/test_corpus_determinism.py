"""Hash-level determinism tests for the full generated corpus.

Task 7 contract — end-to-end invariants that hold across any two
consecutive ``generate-gold`` invocations:

1. **Byte determinism** — every file has the same raw SHA-256.
2. **Content determinism** — every workbook has the same ``content_hash``
   (unpacked zip entries), providing a second line of defense that
   survives zip-envelope changes.
3. **No placeholder leakage** — every ``<foo>`` marker in the templates
   has been resolved by the populator.
4. **Uniqueness** — all 20 workbooks have distinct content hashes.
5. **Corpus-level hash** — a single digest over the 20 (scenario_id,
   content_hash) pairs is stable across runs. Gives PR reviewers a
   one-number "did the corpus change?" answer.
"""

from __future__ import annotations

import hashlib
import re
from pathlib import Path

from openpyxl import load_workbook

from agentic_audit.cli.generate_gold import generate_gold
from agentic_audit.generator.content_hash import content_hash

_REPO_ROOT = Path(__file__).resolve().parents[3]
_MANIFEST_PATH = _REPO_ROOT / "eval" / "gold_scenarios" / "manifest.yaml"

# Detection charset must admit uppercase — a narrow [a-z_0-9] charset
# missed markers like <attribute_A_description> that slipped past the
# populator for the same reason. Keep the test's charset strictly broader
# than the populator's so test can catch populator mistakes.
_PLACEHOLDER_RE = re.compile(r"<\w+>")


# ── byte + content determinism across runs ───────────────────────────


def _byte_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _corpus_byte_hashes(corpus_root: Path) -> dict[str, str]:
    """Byte-hash every .xlsx + .json under corpus_root/tocs/, keyed by filename."""
    tocs = corpus_root / "tocs"
    return {p.name: _byte_hash(p) for p in sorted(tocs.iterdir())}


def _corpus_content_hashes(corpus_root: Path) -> dict[str, str]:
    tocs = corpus_root / "tocs"
    return {p.name: content_hash(p) for p in sorted(tocs.glob("*.xlsx"))}


def test_corpus_byte_hashes_stable_across_regen(tmp_path: Path) -> None:
    first = tmp_path / "first"
    second = tmp_path / "second"
    first.mkdir()
    second.mkdir()

    generate_gold(_MANIFEST_PATH, first)
    generate_gold(_MANIFEST_PATH, second)

    assert _corpus_byte_hashes(first) == _corpus_byte_hashes(second)


def test_corpus_content_hashes_stable_across_regen(tmp_path: Path) -> None:
    first = tmp_path / "first"
    second = tmp_path / "second"
    first.mkdir()
    second.mkdir()

    generate_gold(_MANIFEST_PATH, first)
    generate_gold(_MANIFEST_PATH, second)

    assert _corpus_content_hashes(first) == _corpus_content_hashes(second)


# ── no <placeholder> markers remain ──────────────────────────────────


def test_no_placeholder_markers_in_any_workbook(tmp_path: Path) -> None:
    generate_gold(_MANIFEST_PATH, tmp_path)

    for xlsx_path in sorted((tmp_path / "tocs").glob("*.xlsx")):
        wb = load_workbook(xlsx_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                for value in row:
                    if isinstance(value, str) and _PLACEHOLDER_RE.search(value):
                        raise AssertionError(
                            f"{xlsx_path.name}: sheet {sheet_name!r} has "
                            f"unresolved placeholder in cell value {value!r}"
                        )


# ── content-hash uniqueness ──────────────────────────────────────────


def test_all_twenty_workbooks_have_distinct_content_hashes(tmp_path: Path) -> None:
    """Every scenario must produce a unique workbook — a cheap sanity check
    against accidental fixture collisions (e.g., two scenarios sharing a
    seed and yielding identical output).
    """
    generate_gold(_MANIFEST_PATH, tmp_path)
    hashes = _corpus_content_hashes(tmp_path)
    assert len(hashes) == 20
    assert len(set(hashes.values())) == 20, (
        f"Duplicate content hashes detected across scenarios: {[(n, h) for n, h in hashes.items()]}"
    )


# ── one-number corpus digest ─────────────────────────────────────────


def _corpus_digest(corpus_root: Path) -> str:
    """SHA-256 over ``<name>:<content_hash>\\n`` lines, sorted by name."""
    h = hashlib.sha256()
    for name, ch in sorted(_corpus_content_hashes(corpus_root).items()):
        h.update(f"{name}:{ch}\n".encode())
    return h.hexdigest()


def test_corpus_digest_is_stable_across_regen(tmp_path: Path) -> None:
    first = tmp_path / "first"
    second = tmp_path / "second"
    first.mkdir()
    second.mkdir()
    generate_gold(_MANIFEST_PATH, first)
    generate_gold(_MANIFEST_PATH, second)

    assert _corpus_digest(first) == _corpus_digest(second)
