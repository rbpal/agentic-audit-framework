"""Cross-file consistency tests for DC-9 (Task 13).

Asserts that for every DC-9 scenario, the TOC's "Billing total per
supporting schedule" cell and the billing-calc W/P's "Billing fee"
row either **agree** (pass + non-figure-mismatch exceptions) or
**disagree** (``figure_mismatch``) per the Option-B consistency rules
captured in Q7.13.

This is the invariant the downstream agent will be evaluated on:
detecting the TOC↔W/P contradiction in ``figure_mismatch`` scenarios
while correctly acknowledging the tie-out on every other scenario.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from agentic_audit.cli.generate_gold import generate_gold
from agentic_audit.models import load_manifest

_REPO_ROOT = Path(__file__).resolve().parents[3]
_MANIFEST_PATH = _REPO_ROOT / "eval" / "gold_scenarios" / "manifest.yaml"


def _toc_billing_claim(xlsx_path: Path) -> int:
    """Scan the TOC sheet for the "Billing total per supporting schedule"
    label and return the int from the same row. Raises if not found.
    """
    wb = load_workbook(xlsx_path)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        row_list = list(row)
        for i, cell in enumerate(row_list):
            if isinstance(cell, str) and "Billing total per supporting schedule" in cell:
                for later in row_list[i + 1 :]:
                    if isinstance(later, int):
                        return later
    raise AssertionError(f"{xlsx_path.name} has no billing-total claim row")


def _billing_calc_fee(xlsx_path: Path) -> int:
    """Read the billing-calc W/P's Billing fee (row 6, col B)."""
    wb = load_workbook(xlsx_path)
    ws = wb.active
    fee = ws.cell(row=6, column=2).value
    assert isinstance(fee, int), f"{xlsx_path.name} row 6 col B is not an int: {fee!r}"
    return fee


def test_dc9_pass_scenarios_toc_claim_equals_billing_calc(tmp_path: Path) -> None:
    """Pass scenarios → TOC's claim = billing calc's fee. Auditor correctly
    recorded what the supporting schedule showed.
    """
    generate_gold(_MANIFEST_PATH, tmp_path)
    specs = [
        s
        for s in load_manifest(_MANIFEST_PATH)
        if s.control_id == "DC-9" and s.expected_outcome == "pass"
    ]
    assert len(specs) == 5

    for spec in specs:
        toc_path = tmp_path / "tocs" / f"{spec.scenario_id}_ref.xlsx"
        wp_path = tmp_path / "workpapers" / spec.scenario_id / "billing_calc.xlsx"

        toc_claim = _toc_billing_claim(toc_path)
        wp_fee = _billing_calc_fee(wp_path)

        assert toc_claim == wp_fee, (
            f"{spec.scenario_id} (pass): TOC claim {toc_claim:,} ≠ "
            f"billing calc {wp_fee:,} — should tie for pass scenarios"
        )


def test_dc9_figure_mismatch_scenario_toc_claim_differs_from_billing_calc(
    tmp_path: Path,
) -> None:
    """``figure_mismatch`` scenario → TOC claim ≠ billing calc fee.
    That disagreement IS the cross-file contradiction the agent must detect.
    """
    generate_gold(_MANIFEST_PATH, tmp_path)
    specs = [
        s
        for s in load_manifest(_MANIFEST_PATH)
        if s.control_id == "DC-9" and s.exception_type == "figure_mismatch"
    ]
    assert len(specs) == 1  # q1_exception_dc9_figure_mismatch_01

    for spec in specs:
        toc_path = tmp_path / "tocs" / f"{spec.scenario_id}_ref.xlsx"
        wp_path = tmp_path / "workpapers" / spec.scenario_id / "billing_calc.xlsx"

        toc_claim = _toc_billing_claim(toc_path)
        wp_fee = _billing_calc_fee(wp_path)

        assert toc_claim != wp_fee, (
            f"{spec.scenario_id} (figure_mismatch): TOC claim and billing calc "
            f"BOTH equal {toc_claim:,} — the contradiction rule didn't fire"
        )
        # Sanity: offset should be non-trivial but under 10 % of fee
        diff = abs(toc_claim - wp_fee)
        assert 1_000 <= diff <= wp_fee // 10, (
            f"{spec.scenario_id}: offset {diff:,} is outside the expected "
            f"1k–10% range relative to fee {wp_fee:,}"
        )


def test_dc9_non_figure_mismatch_exceptions_toc_claim_equals_billing_calc(
    tmp_path: Path,
) -> None:
    """DC-9 exceptions OTHER than figure_mismatch → TOC claim = billing calc.

    Signoff_missing, billing_rate_change_without_amendment, boundary_edge_case —
    their defects don't involve billing-total tie-out, so that number should
    still agree across the two files. Task 14 will extend this matrix for
    billing_rate_change_* + governing-doc consistency.
    """
    generate_gold(_MANIFEST_PATH, tmp_path)
    specs = [
        s
        for s in load_manifest(_MANIFEST_PATH)
        if s.control_id == "DC-9"
        and s.expected_outcome == "exception"
        and s.exception_type != "figure_mismatch"
    ]
    assert len(specs) == 4  # 2x signoff_missing, 1x rate_no_amendment, 1x boundary

    for spec in specs:
        toc_path = tmp_path / "tocs" / f"{spec.scenario_id}_ref.xlsx"
        wp_path = tmp_path / "workpapers" / spec.scenario_id / "billing_calc.xlsx"

        toc_claim = _toc_billing_claim(toc_path)
        wp_fee = _billing_calc_fee(wp_path)

        assert toc_claim == wp_fee, (
            f"{spec.scenario_id} ({spec.exception_type}): TOC claim {toc_claim:,} "
            f"≠ billing calc {wp_fee:,} — only figure_mismatch should diverge"
        )


def test_dc2_scenarios_have_no_billing_tieout_cell(tmp_path: Path) -> None:
    """DC-2 (variance_detection) has no billing concept — the TOC must
    NOT emit the billing-tieout cell.
    """
    generate_gold(_MANIFEST_PATH, tmp_path)
    specs = [s for s in load_manifest(_MANIFEST_PATH) if s.control_id == "DC-2"]
    assert len(specs) == 10

    for spec in specs:
        toc_path = tmp_path / "tocs" / f"{spec.scenario_id}_ref.xlsx"
        wb = load_workbook(toc_path)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str) and "Billing total per supporting schedule" in cell:
                    raise AssertionError(
                        f"{spec.scenario_id} (DC-2): unexpectedly emits a "
                        f"billing-tieout cell — should be DC-9-only"
                    )
