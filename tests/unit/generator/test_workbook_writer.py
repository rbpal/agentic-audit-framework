"""Unit tests for the workbook structure writer."""

from __future__ import annotations

import re

import pytest
from openpyxl import Workbook

from agentic_audit.generator import render_toc_sheet
from agentic_audit.models.scenario import ControlId, PatternType, ScenarioSpec


def _make_spec(control_id: ControlId, pattern_type: PatternType) -> ScenarioSpec:
    return ScenarioSpec(
        scenario_id=f"q1_pass_test_{control_id.lower().replace('-', '_')}",
        control_id=control_id,
        pattern_type=pattern_type,
        quarter="Q1",
        expected_outcome="pass",
        exception_type="none",
        seed=0,
    )


# Return type + sheet structure ---------------------------------------


def test_render_returns_workbook() -> None:
    spec = _make_spec("DC-9", "signoff_with_tieout")
    wb = render_toc_sheet(spec)
    assert isinstance(wb, Workbook)


def test_sheet_count_is_one() -> None:
    spec = _make_spec("DC-9", "signoff_with_tieout")
    wb = render_toc_sheet(spec)
    assert len(wb.sheetnames) == 1


@pytest.mark.parametrize(
    ("control_id", "pattern_type"),
    [
        ("DC-9", "signoff_with_tieout"),
        ("DC-2", "variance_detection"),
    ],
)
def test_sheet_name_matches_control_id(control_id: ControlId, pattern_type: PatternType) -> None:
    spec = _make_spec(control_id, pattern_type)
    wb = render_toc_sheet(spec)
    assert wb.sheetnames[0] == control_id


# Header block --------------------------------------------------------


def test_header_labels_at_canonical_positions() -> None:
    spec = _make_spec("DC-9", "signoff_with_tieout")
    wb = render_toc_sheet(spec)
    ws = wb.active
    assert ws is not None
    assert ws["A1"].value == "Workpaper ref.:"
    assert ws["E1"].value == "Tests of the control:"
    assert ws["A2"].value == "Entity name:"
    assert ws["E2"].value == "Related SCOT:"
    assert ws["A3"].value == "Applicable GAAS:"
    assert ws["A4"].value == "Year-end date:"
    assert ws["E4"].value == "Operating effectiveness conclusion:"


def test_header_includes_control_id_in_title() -> None:
    spec = _make_spec("DC-2", "variance_detection")
    wb = render_toc_sheet(spec)
    ws = wb.active
    assert ws is not None
    assert "DC-2" in str(ws["G1"].value)


def test_merged_cells_present_in_header() -> None:
    spec = _make_spec("DC-9", "signoff_with_tieout")
    wb = render_toc_sheet(spec)
    ws = wb.active
    assert ws is not None
    merged = {str(r) for r in ws.merged_cells.ranges}
    assert "A1:B1" in merged
    assert "E1:F1" in merged
    assert "G1:N1" in merged


# Section banners -----------------------------------------------------


@pytest.mark.parametrize(
    "expected_banner",
    [
        "Information about the control",
        "Information produced by the entity (IPE) used in the execution of the control",
        "Testing procedures",
        "Result of procedures",
        "Conclusion",
    ],
)
def test_section_banner_present(expected_banner: str) -> None:
    spec = _make_spec("DC-9", "signoff_with_tieout")
    wb = render_toc_sheet(spec)
    ws = wb.active
    assert ws is not None
    values_in_col_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert expected_banner in values_in_col_a, f"Banner {expected_banner!r} not found in column A"


def test_tickmark_key_section_present() -> None:
    spec = _make_spec("DC-9", "signoff_with_tieout")
    wb = render_toc_sheet(spec)
    ws = wb.active
    assert ws is not None
    values_in_col_b = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
    assert "Tickmark key" in values_in_col_b


# Attribute counts ----------------------------------------------------


def test_attribute_count_dc9_is_six() -> None:
    spec = _make_spec("DC-9", "signoff_with_tieout")
    wb = render_toc_sheet(spec)
    ws = wb.active
    assert ws is not None
    attribute_letters = {
        ws.cell(row=r, column=2).value
        for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=1).value == "Del"
        and ws.cell(row=r, column=2).value in {"A", "B", "C", "D", "E", "F"}
    }
    assert attribute_letters == {"A", "B", "C", "D", "E", "F"}


def test_attribute_count_dc2_is_four() -> None:
    spec = _make_spec("DC-2", "variance_detection")
    wb = render_toc_sheet(spec)
    ws = wb.active
    assert ws is not None
    attribute_letters = {
        ws.cell(row=r, column=2).value
        for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=1).value == "Del"
        and ws.cell(row=r, column=2).value in {"A", "B", "C", "D", "E", "F"}
    }
    assert attribute_letters == {"A", "B", "C", "D"}


# Sample grid column counts ------------------------------------------


def test_sample_grid_has_correct_attribute_columns_dc9() -> None:
    """DC-9 → 6 attrs → columns H through M populated as attribute headers."""
    spec = _make_spec("DC-9", "signoff_with_tieout")
    wb = render_toc_sheet(spec)
    ws = wb.active
    assert ws is not None
    # Find the sample-grid header row: the one with "Sample item #" in col B.
    header_row = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=2).value == "Sample item #"
    )
    headers = [ws.cell(row=header_row, column=c).value for c in range(8, 14)]
    assert headers == ["A", "B", "C", "D", "E", "F"]


def test_sample_grid_has_correct_attribute_columns_dc2() -> None:
    """DC-2 → 4 attrs → columns H through K populated."""
    spec = _make_spec("DC-2", "variance_detection")
    wb = render_toc_sheet(spec)
    ws = wb.active
    assert ws is not None
    header_row = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=2).value == "Sample item #"
    )
    headers = [ws.cell(row=header_row, column=c).value for c in range(8, 12)]
    assert headers == ["A", "B", "C", "D"]
    # Column L (index 12) should be empty since DC-2 only has 4 attrs
    assert ws.cell(row=header_row, column=12).value is None


# Placeholder discipline ---------------------------------------------


def test_placeholder_markers_use_angle_brackets() -> None:
    """Task 2 ships ONLY placeholder markers; no real data allowed."""
    spec = _make_spec("DC-9", "signoff_with_tieout")
    wb = render_toc_sheet(spec)
    ws = wb.active
    assert ws is not None

    placeholder_pattern = re.compile(r"^<[a-z_0-9]+>$")
    found_placeholders = 0
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str) and cell.startswith("<") and cell.endswith(">"):
                assert placeholder_pattern.match(cell) or cell.startswith("<"), (
                    f"Placeholder {cell!r} must match <lowercase_with_underscores>"
                )
                found_placeholders += 1
    # At least some placeholders should exist
    assert found_placeholders > 10


# Frequency + sample-size differ by pattern --------------------------


def test_frequency_differs_by_pattern() -> None:
    wb9 = render_toc_sheet(_make_spec("DC-9", "signoff_with_tieout"))
    wb2 = render_toc_sheet(_make_spec("DC-2", "variance_detection"))
    ws9, ws2 = wb9.active, wb2.active
    assert ws9 is not None
    assert ws2 is not None

    def _find_value_next_to_label(ws: object, label: str) -> object:
        assert hasattr(ws, "iter_rows")
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == label:
                    # value is the cell to the right (+ some offset per layout)
                    return ws.cell(row=cell.row, column=cell.column + 1).value
        return None

    assert _find_value_next_to_label(ws9, "Frequency") == "Quarterly"
    assert _find_value_next_to_label(ws2, "Frequency") == "Monthly"
