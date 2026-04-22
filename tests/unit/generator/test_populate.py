"""Integration tests for populate_workbook.

Each test renders a Task-2 workbook, populates it, and asserts on the
resulting workbook. No filesystem I/O — all in memory.
"""

from __future__ import annotations

import datetime as dt
import random
import re

import pytest

from agentic_audit.generator import populate_workbook, render_toc_sheet
from agentic_audit.generator.populate import _resolve_placeholder
from agentic_audit.models.scenario import ControlId, PatternType, ScenarioSpec

_PLACEHOLDER_RE = re.compile(r"^<[a-z_0-9]+>$")


def _make_spec(
    control_id: ControlId = "DC-9",
    pattern_type: PatternType = "signoff_with_tieout",
    quarter: str = "Q1",
    expected_outcome: str = "pass",
    exception_type: str = "none",
    seed: int = 42,
) -> ScenarioSpec:
    return ScenarioSpec(
        scenario_id=f"x_populate_test_{seed}",
        control_id=control_id,
        pattern_type=pattern_type,
        quarter=quarter,  # type: ignore[arg-type]
        expected_outcome=expected_outcome,  # type: ignore[arg-type]
        exception_type=exception_type,  # type: ignore[arg-type]
        seed=seed,
    )


def _canonical_dump(wb: object) -> list[tuple[int, int, object]]:
    """Stable (row, col, value) triples for cell-by-cell comparison."""
    assert hasattr(wb, "active")
    ws = wb.active  # type: ignore[attr-defined]
    return [
        (cell.row, cell.column, cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    ]


# ── Placeholder removal ───────────────────────────────────────────────


def test_populate_removes_all_placeholders_dc9() -> None:
    spec = _make_spec("DC-9", "signoff_with_tieout")
    wb = populate_workbook(render_toc_sheet(spec), spec)
    ws = wb.active
    assert ws is not None
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str):
                assert not _PLACEHOLDER_RE.match(cell), (
                    f"Placeholder {cell!r} remains after populate"
                )


def test_populate_removes_all_placeholders_dc2() -> None:
    spec = _make_spec("DC-2", "variance_detection")
    wb = populate_workbook(render_toc_sheet(spec), spec)
    ws = wb.active
    assert ws is not None
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str):
                assert not _PLACEHOLDER_RE.match(cell), (
                    f"Placeholder {cell!r} remains after populate"
                )


# ── Seed reproducibility ─────────────────────────────────────────────


def test_same_seed_identical_workbooks() -> None:
    spec = _make_spec(seed=42)
    wb1 = populate_workbook(render_toc_sheet(spec), spec)
    wb2 = populate_workbook(render_toc_sheet(spec), spec)
    assert _canonical_dump(wb1) == _canonical_dump(wb2)


def test_different_seeds_produce_different_workbooks() -> None:
    spec1 = _make_spec(seed=42)
    spec2 = _make_spec(seed=43)
    wb1 = populate_workbook(render_toc_sheet(spec1), spec1)
    wb2 = populate_workbook(render_toc_sheet(spec2), spec2)
    # Cell-by-cell dumps must differ on at least one value
    assert _canonical_dump(wb1) != _canonical_dump(wb2)


def test_populate_is_idempotent() -> None:
    """Running populate_workbook twice on the same wb should be stable.

    After first call no placeholders remain; second call is a no-op.
    """
    spec = _make_spec(seed=42)
    wb = populate_workbook(render_toc_sheet(spec), spec)
    dump_after_one = _canonical_dump(wb)
    wb = populate_workbook(wb, spec)
    assert _canonical_dump(wb) == dump_after_one


# ── No global random pollution ────────────────────────────────────────


def test_no_global_random_state_pollution() -> None:
    random.seed(99)
    before = [random.random() for _ in range(5)]
    random.seed(99)
    spec = _make_spec(seed=42)
    populate_workbook(render_toc_sheet(spec), spec)
    after = [random.random() for _ in range(5)]
    assert before == after


# ── Date-in-quarter ───────────────────────────────────────────────────


@pytest.mark.parametrize("quarter,month_range", [("Q1", (1, 3)), ("Q3", (7, 9))])
def test_dates_fall_in_declared_quarter(quarter: str, month_range: tuple[int, int]) -> None:
    spec = _make_spec(quarter=quarter)
    wb = populate_workbook(render_toc_sheet(spec), spec)
    ws = wb.active
    assert ws is not None
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, dt.date):
                assert month_range[0] <= cell.month <= month_range[1], (
                    f"{quarter} date {cell} outside month range {month_range}"
                )


# ── Currency discipline (§4.5: USD only) ─────────────────────────────


@pytest.mark.parametrize("currency_symbol", ["£", "GBP", "€"])
def test_no_gbp_or_eur_in_generated_text(currency_symbol: str) -> None:
    spec = _make_spec(seed=42)
    wb = populate_workbook(render_toc_sheet(spec), spec)
    ws = wb.active
    assert ws is not None
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str):
                assert currency_symbol not in cell, (
                    f"Currency {currency_symbol} found in cell: {cell!r}"
                )


# ── Outcome-aware population ─────────────────────────────────────────


def test_pass_scenario_all_tickmarks_are_a() -> None:
    spec = _make_spec(
        control_id="DC-9",
        pattern_type="signoff_with_tieout",
        expected_outcome="pass",
    )
    wb = populate_workbook(render_toc_sheet(spec), spec)
    ws = wb.active
    assert ws is not None
    # Find the sample-grid region by locating the "Sample item #" header row
    header_row = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=2).value == "Sample item #"
    )
    # Check tickmarks in rows following the header (columns H-M = 8-13)
    for sample_row in range(header_row + 1, header_row + 3):  # 2 sample rows
        for col in range(8, 14):
            value = ws.cell(row=sample_row, column=col).value
            if value is not None:
                assert value == "a", f"Pass-scenario tickmark at ({sample_row}, {col}) is {value!r}"


def test_exception_scenario_has_one_x_per_sample() -> None:
    spec = _make_spec(
        control_id="DC-9",
        pattern_type="signoff_with_tieout",
        expected_outcome="exception",
        exception_type="figure_mismatch",  # maps to attribute D
    )
    wb = populate_workbook(render_toc_sheet(spec), spec)
    ws = wb.active
    assert ws is not None
    header_row = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=2).value == "Sample item #"
    )
    # "D" is at column H + 3 = 11 (K)
    for sample_row in range(header_row + 1, header_row + 3):
        d_value = ws.cell(row=sample_row, column=11).value
        assert d_value == "X", f"Expected X at row {sample_row} col K, got {d_value!r}"
        # Other attributes remain "a"
        for col in (8, 9, 10, 12, 13):
            value = ws.cell(row=sample_row, column=col).value
            if value is not None:
                assert value == "a", f"Non-failing attribute at ({sample_row}, {col}) is {value!r}"


def test_pass_scenario_has_effectiveness_effective() -> None:
    spec = _make_spec(expected_outcome="pass")
    wb = populate_workbook(render_toc_sheet(spec), spec)
    ws = wb.active
    assert ws is not None
    # Effectiveness conclusion appears in the header row 4 and in the conclusion
    effectiveness_values = [
        ws.cell(row=r, column=c).value
        for r in range(1, ws.max_row + 1)
        for c in range(1, 15)
        if ws.cell(row=r, column=c).value in {"Effective", "Not effective"}
    ]
    assert all(v == "Effective" for v in effectiveness_values)
    assert len(effectiveness_values) >= 1


def test_exception_scenario_has_effectiveness_not_effective() -> None:
    spec = _make_spec(expected_outcome="exception", exception_type="figure_mismatch")
    wb = populate_workbook(render_toc_sheet(spec), spec)
    ws = wb.active
    assert ws is not None
    effectiveness_values = [
        ws.cell(row=r, column=c).value
        for r in range(1, ws.max_row + 1)
        for c in range(1, 15)
        if ws.cell(row=r, column=c).value in {"Effective", "Not effective"}
    ]
    assert all(v == "Not effective" for v in effectiveness_values)


# ── Error-path tests ──────────────────────────────────────────────────


def test_unrecognized_placeholder_raises() -> None:
    """If _resolve_placeholder ever sees an unknown name, it must raise."""
    spec = _make_spec()

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    fake_cell = ws["A1"]
    with pytest.raises(ValueError, match="Unrecognized placeholder"):
        _resolve_placeholder("totally_made_up", fake_cell, random.Random(0), spec)


def test_unknown_sample_kind_raises() -> None:
    spec = _make_spec()
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    fake_cell = ws["A1"]
    with pytest.raises(ValueError, match="Unknown sample placeholder kind"):
        _resolve_placeholder("sample_1_bogus", fake_cell, random.Random(0), spec)


def test_unknown_attribute_kind_raises() -> None:
    spec = _make_spec()
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    fake_cell = ws["A1"]
    with pytest.raises(ValueError, match="Unknown attribute placeholder kind"):
        _resolve_placeholder("attribute_A_bogus", fake_cell, random.Random(0), spec)
