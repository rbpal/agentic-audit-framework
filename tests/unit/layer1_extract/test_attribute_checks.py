"""Behavior tests for `attribute_checks.check_attribute`.

Loads real v2 workpaper xlsx files into `BronzeWorkpaperRow` lists
(the same shape `BronzeReader.read` returns at runtime) and runs the
10 dispatch branches against them.

Coverage matrix:

- 1 happy-path test per (control, attribute) using a quarter where the
  TOC says the attribute should `pass`.
- 1 fail test per attribute. Where the corpus has a real failing
  quarter (DC-2.B Q4, DC-2.C Q3, DC-9.D Q4), use it. Otherwise mutate
  a passing fixture to break the invariant.
- DC-9.D Q1 specifically asserts `n/a` (no prior period).
- 2 dispatch error tests retained from the stub suite (DC-2 + E,
  DC-2 + F).
- 1 cross-cutting `n/a` assertion that DC-9.D Q1 is the only `n/a`
  the current corpus produces.

Workpaper xlsx live at `eval/gold_scenarios/workpapers/`.
"""

from __future__ import annotations

import copy
from datetime import UTC, datetime
from pathlib import Path

import openpyxl
import pytest

from agentic_audit.layer1_extract.attribute_checks import check_attribute
from agentic_audit.layer1_extract.bronze_reader import BronzeWorkpaperRow

CORPUS = Path("eval/gold_scenarios/workpapers")
UTC_TS = datetime(2026, 5, 1, 12, 0, 0, tzinfo=UTC)


def _load_rows(xlsx_path: Path) -> list[BronzeWorkpaperRow]:
    """Read an xlsx into the same row shape `BronzeReader.read` returns."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out: list[BronzeWorkpaperRow] = []
    # Filename like `dc9_Q1_ref.xlsx` → control DC-9, quarter Q1.
    name = xlsx_path.stem  # e.g. 'dc9_Q1_ref'
    parts = name.split("_")
    control_id = f"DC-{parts[0][2:]}"  # 'dc9' → 'DC-9'
    quarter = parts[1]  # 'Q1'
    file_hash = "f" * 64  # synthetic; not used by checks
    source_path = str(xlsx_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            non_empty = {f"col_{j:02d}": str(v) for j, v in enumerate(row) if v is not None}
            if not non_empty:
                continue
            out.append(
                BronzeWorkpaperRow(
                    source_path=source_path,
                    file_hash=file_hash,
                    engagement_id="alpha-pension-fund-2025",
                    control_id=control_id,  # type: ignore[arg-type]
                    quarter=quarter,  # type: ignore[arg-type]
                    sheet_name=sheet_name,
                    row_index=row_idx,
                    raw_data=non_empty,
                    ingested_at=UTC_TS,
                    ingested_by="test",
                )
            )
    wb.close()
    return out


def _set_cell(
    rows: list[BronzeWorkpaperRow],
    sheet: str,
    row_index: int,
    col_key: str,
    value: str | None,
) -> list[BronzeWorkpaperRow]:
    """Return a deep copy of `rows` with one cell rewritten (or removed
    if `value` is None). The original list is untouched."""
    copied = copy.deepcopy(rows)
    for r in copied:
        if r.sheet_name == sheet and r.row_index == row_index:
            if value is None:
                r.raw_data.pop(col_key, None)
            else:
                r.raw_data[col_key] = value
            break
    return copied


# ---------- DC-2.A — data-load tie-out -----------------------------------


def test_dc2_a_pass_q1() -> None:
    rows = _load_rows(CORPUS / "dc2_Q1_ref.xlsx")
    result = check_attribute("DC-2", "A", rows, toc=None)
    assert result.status == "pass"
    assert result.extracted_value == {"feed_total": 18665242.0, "workbook_total": 18665242.0}


def test_dc2_a_fail_when_workbook_does_not_tie() -> None:
    rows = _load_rows(CORPUS / "dc2_Q1_ref.xlsx")
    rows = _set_cell(rows, "DC-2 Variance", 7, "col_01", "99999999")
    result = check_attribute("DC-2", "A", rows, toc=None)
    assert result.status == "fail"
    assert "does not tie" in (result.notes or "")


# ---------- DC-2.B — variance has explanation ----------------------------


def test_dc2_b_pass_q1() -> None:
    rows = _load_rows(CORPUS / "dc2_Q1_ref.xlsx")
    result = check_attribute("DC-2", "B", rows, toc=None)
    assert result.status == "pass"


def test_dc2_b_fail_q4_blank_explanation() -> None:
    """Real corpus failure: dc2_Q4 row 10 has Above threshold = Yes
    but Explanation cell is blank."""
    rows = _load_rows(CORPUS / "dc2_Q4_ref.xlsx")
    result = check_attribute("DC-2", "B", rows, toc=None)
    assert result.status == "fail"
    assert "missing Explanation" in (result.notes or "")
    assert 10 in (result.extracted_value or {}).get("rows_missing_explanation", [])


# ---------- DC-2.C — variance has source tie -----------------------------


def test_dc2_c_pass_q1() -> None:
    rows = _load_rows(CORPUS / "dc2_Q1_ref.xlsx")
    result = check_attribute("DC-2", "C", rows, toc=None)
    assert result.status == "pass"


def test_dc2_c_fail_q3_source_tie_no() -> None:
    """Real corpus failure: dc2_Q3 has 4 above-threshold rows with
    Source tie != 'Yes'."""
    rows = _load_rows(CORPUS / "dc2_Q3_ref.xlsx")
    result = check_attribute("DC-2", "C", rows, toc=None)
    assert result.status == "fail"
    bad = (result.extracted_value or {}).get("rows_with_source_tie_no", [])
    assert sorted(bad) == [10, 11, 13, 14]


# ---------- DC-2.D — reviewer sign-off -----------------------------------


def test_dc2_d_pass_q1() -> None:
    rows = _load_rows(CORPUS / "dc2_Q1_ref.xlsx")
    result = check_attribute("DC-2", "D", rows, toc=None)
    assert result.status == "pass"
    assert " — " in str(result.extracted_value)


def test_dc2_d_fail_when_reviewer_cell_empty() -> None:
    rows = _load_rows(CORPUS / "dc2_Q1_ref.xlsx")
    rows = _set_cell(rows, "DC-2 Variance", 17, "col_01", None)
    result = check_attribute("DC-2", "D", rows, toc=None)
    assert result.status == "fail"


# ---------- DC-9.A — preparer sign-off ----------------------------------


def test_dc9_a_pass_q1() -> None:
    rows = _load_rows(CORPUS / "dc9_Q1_ref.xlsx")
    result = check_attribute("DC-9", "A", rows, toc=None)
    assert result.status == "pass"
    assert "FV" in str(result.extracted_value)


def test_dc9_a_fail_when_preparer_cell_corrupt() -> None:
    rows = _load_rows(CORPUS / "dc9_Q1_ref.xlsx")
    rows = _set_cell(rows, "DC-9 Billing", 4, "col_01", "TBD")  # no '—'
    result = check_attribute("DC-9", "A", rows, toc=None)
    assert result.status == "fail"


# ---------- DC-9.B — reviewer sign-off ----------------------------------


def test_dc9_b_pass_q1() -> None:
    rows = _load_rows(CORPUS / "dc9_Q1_ref.xlsx")
    result = check_attribute("DC-9", "B", rows, toc=None)
    assert result.status == "pass"
    assert "DM" in str(result.extracted_value)


def test_dc9_b_fail_when_reviewer_cell_missing() -> None:
    rows = _load_rows(CORPUS / "dc9_Q1_ref.xlsx")
    rows = _set_cell(rows, "DC-9 Billing", 5, "col_01", None)
    result = check_attribute("DC-9", "B", rows, toc=None)
    assert result.status == "fail"


# ---------- DC-9.C — billing-fee math tie -------------------------------


def test_dc9_c_pass_q1() -> None:
    rows = _load_rows(CORPUS / "dc9_Q1_ref.xlsx")
    result = check_attribute("DC-9", "C", rows, toc=None)
    assert result.status == "pass"
    ev = result.extracted_value or {}
    assert ev["asset"] == 337995108.0
    assert ev["rate"] == pytest.approx(0.0025)


def test_dc9_c_pass_q3_internal_tie_holds_even_though_toc_says_fail() -> None:
    """DC-9 Q3 has a TOC-vs-W/P cross-file figure mismatch, but the W/P
    internal billing math (asset × rate ≈ fee) ties. Layer 1 (single-file)
    correctly returns pass; cross-file detection lives in
    silver.cross_file_validations. Documented in step_04 doc § task_03.1."""
    rows = _load_rows(CORPUS / "dc9_Q3_ref.xlsx")
    result = check_attribute("DC-9", "C", rows, toc=None)
    assert result.status == "pass"


def test_dc9_c_fail_when_fee_does_not_tie_to_asset_times_rate() -> None:
    rows = _load_rows(CORPUS / "dc9_Q1_ref.xlsx")
    rows = _set_cell(rows, "DC-9 Billing", 10, "col_01", "100")  # nonsense fee
    result = check_attribute("DC-9", "C", rows, toc=None)
    assert result.status == "fail"
    assert "does not tie" in (result.notes or "")


# ---------- DC-9.D — rate change with amendment -------------------------


def test_dc9_d_q1_returns_n_a() -> None:
    """Q1 has no prior period → n/a."""
    rows = _load_rows(CORPUS / "dc9_Q1_ref.xlsx")
    result = check_attribute("DC-9", "D", rows, toc=None)
    assert result.status == "n/a"
    assert "no prior period" in (result.notes or "")


def test_dc9_d_q3_pass_no_rate_change() -> None:
    """Q3 has prior=current=0.50%; no change ⇒ no amendment needed."""
    rows = _load_rows(CORPUS / "dc9_Q3_ref.xlsx")
    result = check_attribute("DC-9", "D", rows, toc=None)
    assert result.status == "pass"
    ev = result.extracted_value or {}
    assert ev["rate_change"] is False


def test_dc9_d_q4_fail_no_amendment_filed() -> None:
    """Q4 real corpus failure: prior=0.50%, current=0.75%, amendment
    cell reads 'NO AMENDMENT FILED'."""
    rows = _load_rows(CORPUS / "dc9_Q4_ref.xlsx")
    result = check_attribute("DC-9", "D", rows, toc=None)
    assert result.status == "fail"
    ev = result.extracted_value or {}
    assert ev["prior_rate"] == pytest.approx(0.005)
    assert ev["current_rate"] == pytest.approx(0.0075)
    assert "NO AMENDMENT" in ev["amendment"].upper()


def test_dc9_d_pass_when_rate_changed_and_amendment_present() -> None:
    """Synthesise a Q4 row where the amendment is properly filed."""
    rows = _load_rows(CORPUS / "dc9_Q4_ref.xlsx")
    rows = _set_cell(rows, "DC-9 Billing", 16, "col_01", "Amendment 2025-10-26 (signed by CFO)")
    result = check_attribute("DC-9", "D", rows, toc=None)
    assert result.status == "pass"


# ---------- DC-9.E — asset roll-forward ----------------------------------


def test_dc9_e_pass_q1() -> None:
    rows = _load_rows(CORPUS / "dc9_Q1_ref.xlsx")
    result = check_attribute("DC-9", "E", rows, toc=None)
    assert result.status == "pass"


def test_dc9_e_fail_when_closing_does_not_tie() -> None:
    rows = _load_rows(CORPUS / "dc9_Q1_ref.xlsx")
    rows = _set_cell(rows, "DC-9 Billing", 22, "col_01", "999999999")
    result = check_attribute("DC-9", "E", rows, toc=None)
    assert result.status == "fail"
    assert "does not tie" in (result.notes or "")


# ---------- DC-9.F — ownership share -------------------------------------


def test_dc9_f_pass_q1() -> None:
    rows = _load_rows(CORPUS / "dc9_Q1_ref.xlsx")
    result = check_attribute("DC-9", "F", rows, toc=None)
    assert result.status == "pass"
    ev = result.extracted_value or {}
    assert ev["total"] == 100.0


def test_dc9_f_fail_when_pcts_do_not_sum_to_100() -> None:
    rows = _load_rows(CORPUS / "dc9_Q1_ref.xlsx")
    rows = _set_cell(rows, "DC-9 Billing", 26, "col_02", "50")  # was 40; total now 110
    result = check_attribute("DC-9", "F", rows, toc=None)
    assert result.status == "fail"
    assert "expected 100" in (result.notes or "")


# ---------- dispatch error paths ----------------------------------------


def test_dc2_attribute_e_raises_key_error() -> None:
    """DC-2 only defines A-D; E is a contract violation."""
    with pytest.raises(KeyError, match="DC-2 does not define attribute_id=E"):
        check_attribute("DC-2", "E", rows=[], toc=None)  # type: ignore[arg-type]


def test_dc2_attribute_f_raises_key_error() -> None:
    with pytest.raises(KeyError, match="DC-2 does not define attribute_id=F"):
        check_attribute("DC-2", "F", rows=[], toc=None)  # type: ignore[arg-type]


# ---------- full sweep — gold answers cross-check -----------------------


@pytest.mark.parametrize(
    ("control", "quarter", "expected"),
    [
        ("DC-2", "Q1", {"A": "pass", "B": "pass", "C": "pass", "D": "pass"}),
        ("DC-2", "Q2", {"A": "pass", "B": "pass", "C": "pass", "D": "pass"}),
        ("DC-2", "Q3", {"A": "pass", "B": "pass", "C": "fail", "D": "pass"}),
        ("DC-2", "Q4", {"A": "pass", "B": "fail", "C": "pass", "D": "pass"}),
        # DC-9 Q3 attribute C: Layer 1 single-file says pass (W/P math ties).
        # TOC's gold says fail because of cross-file mismatch — that's
        # silver.cross_file_validations territory.
        (
            "DC-9",
            "Q1",
            {"A": "pass", "B": "pass", "C": "pass", "D": "n/a", "E": "pass", "F": "pass"},
        ),
        (
            "DC-9",
            "Q2",
            {"A": "pass", "B": "pass", "C": "pass", "D": "pass", "E": "pass", "F": "pass"},
        ),
        (
            "DC-9",
            "Q3",
            {"A": "pass", "B": "pass", "C": "pass", "D": "pass", "E": "pass", "F": "pass"},
        ),
        (
            "DC-9",
            "Q4",
            {"A": "pass", "B": "pass", "C": "pass", "D": "fail", "E": "pass", "F": "pass"},
        ),
    ],
)
def test_full_sweep_matches_layer1_expected_results(
    control: str,
    quarter: str,
    expected: dict[str, str],
) -> None:
    rows = _load_rows(CORPUS / f"dc{control[3]}_{quarter}_ref.xlsx")
    actual = {
        attr: check_attribute(control, attr, rows, toc=None).status  # type: ignore[arg-type]
        for attr in expected
    }
    assert actual == expected, f"({control}, {quarter}) mismatch: got {actual}, expected {expected}"
