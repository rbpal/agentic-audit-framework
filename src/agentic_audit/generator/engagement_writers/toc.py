"""Engagement TOC — one file, 2 sheets (DC-2, DC-9), kept simple.

Each sheet is a tabular tickmark matrix: rows = attributes, columns =
quarters (Q1/Q2/Q3/Q4). Cells carry ``"a"`` (satisfied) or ``"X"``
(exception). A brief conclusion row follows the tickmark matrix.

The DC-9 sheet additionally emits a per-quarter *billing-fee claim*
row. For pass / non-``figure_mismatch`` quarters the claim equals
the canonical W/P billing fee; for Q3 (when declared
``dc9_figure_mismatch``) the claim intentionally diverges by ~5 % —
that's the cross-file contradiction the agent must detect.

v2's TOC does NOT replicate v1's banner rows, IPE block, multi-instance
block, sample grid, etc. Detail evidence lives in the per-quarter W/Ps.
"""

from __future__ import annotations

from openpyxl import Workbook

from agentic_audit.generator.engagement_writers.common import canonical_billing_fee
from agentic_audit.models.engagement import (
    ControlId,
    EngagementSpec,
    Quarter,
    QuarterDefect,
    quarter_control,
)

# ── Attribute descriptions (what each tickmark means) ───────────────

_DC2_ATTRIBUTES: tuple[tuple[str, str], ...] = (
    ("A", "Current-period accrual data loaded completely"),
    ("B", "Variances above threshold have recorded explanation"),
    ("C", "Explanations are consistent with upstream source"),
    ("D", "Reviewer signed off on the variance analysis"),
)

_DC9_ATTRIBUTES: tuple[tuple[str, str], ...] = (
    ("A", "Preparer signed off on the Checklist"),
    ("B", "Independent reviewer signed off"),
    ("C", "Billing formulas tie to underlying supporting schedule"),
    ("D", "Billing rate change supported by governing-document amendment"),
    ("E", "Asset additions and retirements on the supporting schedule"),
    ("F", "Ownership-share percentages match supporting reference file"),
)

# ── Defect → failing-attribute mapping ──────────────────────────────

_DEFECT_TO_ATTRIBUTE: dict[QuarterDefect, str | None] = {
    "none": None,
    "dc9_figure_mismatch": "C",
    "dc9_rate_change_with_amendment": None,  # benign — all attrs pass
    "dc9_rate_change_without_amendment": "D",
    "dc2_variance_no_explanation": "B",
    "dc2_variance_explanation_inadequate": "C",
    "dc2_variance_boundary": "A",
}


def _tickmark(defect: QuarterDefect, attribute: str) -> str:
    """``"X"`` if ``defect`` targets ``attribute``, else ``"a"``."""
    failing = _DEFECT_TO_ATTRIBUTE[defect]
    return "X" if failing == attribute else "a"


def _toc_billing_fee_claim(spec: EngagementSpec, quarter: Quarter) -> int:
    """TOC's *asserted* billing fee for a DC-9 quarter.

    Pass / non-figure_mismatch → matches the canonical W/P fee.
    ``dc9_figure_mismatch`` → offset by ~5 % (rounded to nearest 1 k,
    min 1 k) so the TOC disagrees with the W/P.
    """
    qc = quarter_control(spec, "DC-9", quarter)
    canonical = canonical_billing_fee(spec, quarter)
    if qc.defect == "dc9_figure_mismatch":
        offset = max(1_000, round(canonical * 0.05 / 1_000) * 1_000)
        return canonical + offset
    return canonical


def _emit_control_sheet(
    wb: Workbook,
    spec: EngagementSpec,
    control_id: ControlId,
    attributes: tuple[tuple[str, str], ...],
) -> None:
    """Shared layout for the DC-2 and DC-9 sheets.

    Tabular: row per attribute × column per quarter tickmark.
    """
    ws = wb.create_sheet(title=control_id)

    ws.cell(row=1, column=1, value="Entity")
    ws.cell(row=1, column=2, value=spec.entity_name)
    ws.cell(row=1, column=4, value="Control")
    ws.cell(row=1, column=5, value=control_id)
    ws.cell(row=2, column=1, value="Period of reliance")
    ws.cell(row=2, column=2, value=f"{spec.year} (Q1–Q4)")

    # Row 4 — column headers
    ws.cell(row=4, column=1, value="Attribute")
    ws.cell(row=4, column=2, value="Description")
    for i, quarter in enumerate(("Q1", "Q2", "Q3", "Q4"), start=3):
        ws.cell(row=4, column=i, value=quarter)

    # Rows 5..N — attribute rows with per-quarter tickmarks
    fail_quarters: list[str] = []
    for i, (letter, description) in enumerate(attributes, start=5):
        ws.cell(row=i, column=1, value=letter)
        ws.cell(row=i, column=2, value=description)
        for j, quarter in enumerate(("Q1", "Q2", "Q3", "Q4"), start=3):
            qc = quarter_control(spec, control_id, quarter)  # type: ignore[arg-type]
            mark = _tickmark(qc.defect, letter)
            ws.cell(row=i, column=j, value=mark)
            if mark == "X" and quarter not in fail_quarters:
                fail_quarters.append(quarter)

    # Conclusion — row after last attribute + 2 blank rows
    conclusion_row = 5 + len(attributes) + 2
    verdict = "Not effective" if fail_quarters else "Effective"
    # Sort so the "Quarters with exceptions" list reads chronologically
    # regardless of which attribute's row encountered the X first.
    fail_quarters_sorted = sorted(fail_quarters, key=lambda q: ("Q1", "Q2", "Q3", "Q4").index(q))
    ws.cell(row=conclusion_row, column=1, value="Overall conclusion")
    ws.cell(row=conclusion_row, column=2, value=verdict)
    ws.cell(row=conclusion_row + 1, column=1, value="Quarters with exceptions")
    ws.cell(
        row=conclusion_row + 1,
        column=2,
        value=", ".join(fail_quarters_sorted) if fail_quarters_sorted else "None",
    )

    # Legend — two rows further down
    legend_row = conclusion_row + 3
    ws.cell(row=legend_row, column=1, value="Legend")
    ws.cell(row=legend_row + 1, column=1, value="a")
    ws.cell(row=legend_row + 1, column=2, value="Attribute satisfied")
    ws.cell(row=legend_row + 2, column=1, value="X")
    ws.cell(row=legend_row + 2, column=2, value="Attribute failed / exception noted")


def _emit_dc9_billing_claim_row(wb: Workbook, spec: EngagementSpec) -> None:
    """Add the cross-file billing claim row to the DC-9 sheet.

    Placed 2 rows below the tickmark matrix. Q3 value diverges from
    the W/P for ``dc9_figure_mismatch`` — that's the cross-file defect.
    """
    ws = wb["DC-9"]
    # Tickmark rows occupy 5..10 (6 attributes). Claim row at 12.
    claim_row = 12
    ws.cell(row=claim_row, column=1, value="Billing fee per W/P (USD)")
    ws.cell(row=claim_row, column=2, value="(TOC claim)")
    for j, quarter in enumerate(("Q1", "Q2", "Q3", "Q4"), start=3):
        ws.cell(row=claim_row, column=j, value=_toc_billing_fee_claim(spec, quarter))  # type: ignore[arg-type]


def render_engagement_toc(spec: EngagementSpec) -> Workbook:
    """Emit the engagement TOC — one file, two sheets (DC-2, DC-9).

    Byte-deterministic per ``spec.seed`` (tickmarks + billing claim values
    are derived from the engagement's declared defects and canonical
    billing, no rng involvement).
    """
    wb = Workbook()
    # Workbook() has a default empty sheet we don't want — remove it.
    default = wb.active
    assert default is not None
    wb.remove(default)

    _emit_control_sheet(wb, spec, "DC-2", _DC2_ATTRIBUTES)
    _emit_control_sheet(wb, spec, "DC-9", _DC9_ATTRIBUTES)
    _emit_dc9_billing_claim_row(wb, spec)

    return wb
