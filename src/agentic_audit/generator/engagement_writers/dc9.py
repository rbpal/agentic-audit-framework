"""DC-9 quarterly workpaper — all 6 attributes in one worksheet.

One file per quarter. Each file contains all the evidence needed to
approve or deny every DC-9 attribute (A–F) for its quarter.

Section layout (single worksheet, "DC-9 Billing"):

* Rows 1–3  — Header (entity, quarter, as-of date)
* Row  4    — Attribute A : preparer initials + date
* Row  5    — Attribute B : reviewer initials + date
* Rows 7–10 — Attribute C : asset value, billing rate, billing fee
* Rows 12–15— Attribute D : prior rate, current rate, change date, amendment
* Rows 17–21— Attribute E : asset roll-forward (opening / additions /
                            retirements / closing, tying by construction)
* Rows 23–27— Attribute F : LP ownership share table (sums to 100 %)

Defect expression in this writer:

* ``dc9_rate_change_with_amendment``   → row 15 shows the amendment ref;
                                          attribute D passes.
* ``dc9_rate_change_without_amendment``→ row 15 says "NO AMENDMENT FILED";
                                          attribute D fails.
* ``dc9_figure_mismatch``               → lives in the TOC writer, not
                                          here. This W/P always emits the
                                          *canonical* billing fee; the TOC
                                          claims a different value in Q3,
                                          producing the cross-file defect.

Attributes A, B, E, F always pass in the §5 plan — this writer emits
clean, internally-consistent data for them.
"""

from __future__ import annotations

from openpyxl import Workbook

from agentic_audit.generator.engagement_writers.common import (
    billing_rate,
    canonical_asset_value,
    canonical_billing_fee,
    date_in_quarter,
    pick_initials,
    prior_quarter,
    quarter_end_date,
    rate_changed_this_quarter,
    seeded_rng,
)
from agentic_audit.models.engagement import EngagementSpec, Quarter, quarter_control

_OWNERSHIP_LPS: tuple[tuple[str, int], ...] = (
    ("Alpha Institutional LP", 40),
    ("Beta Family Office LP", 30),
    ("Gamma Endowment LP", 30),
)


def render_dc9_quarter(spec: EngagementSpec, quarter: Quarter) -> Workbook:
    """Emit the DC-9 W/P for one quarter. Byte-deterministic per spec.seed.

    The returned ``Workbook`` has a single sheet named ``"DC-9 Billing"``.
    Callers are responsible for ``wb.save(path)``.
    """
    qc = quarter_control(spec, "DC-9", quarter)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "DC-9 Billing"

    # ── Rows 1–3: header ────────────────────────────────────────────
    ws.cell(row=1, column=1, value="Entity")
    ws.cell(row=1, column=2, value=spec.entity_name)
    ws.cell(row=2, column=1, value="Quarter")
    ws.cell(row=2, column=2, value=f"{quarter} {spec.year}")
    ws.cell(row=3, column=1, value="As-of date")
    ws.cell(row=3, column=2, value=quarter_end_date(quarter).isoformat())

    # ── Rows 4–5: sign-offs (attributes A + B) ──────────────────────
    prep_rng = seeded_rng(spec, f"dc9_{quarter}_preparer")
    prep_initials = pick_initials(prep_rng)
    prep_date = date_in_quarter(prep_rng, quarter)
    ws.cell(row=4, column=1, value="Preparer (Attribute A)")
    ws.cell(row=4, column=2, value=f"{prep_initials} — {prep_date.isoformat()}")

    rev_rng = seeded_rng(spec, f"dc9_{quarter}_reviewer")
    rev_initials = pick_initials(rev_rng)
    rev_date = date_in_quarter(rev_rng, quarter)
    ws.cell(row=5, column=1, value="Reviewer (Attribute B)")
    ws.cell(row=5, column=2, value=f"{rev_initials} — {rev_date.isoformat()}")

    # ── Rows 7–10: billing calculation (attribute C) ────────────────
    asset_value = canonical_asset_value(spec, quarter)
    rate_str = billing_rate(quarter)
    billing_fee = canonical_billing_fee(spec, quarter)
    ws.cell(row=7, column=1, value="Billing calculation (Attribute C)")
    ws.cell(row=8, column=1, value="Asset value (USD)")
    ws.cell(row=8, column=2, value=asset_value)
    ws.cell(row=9, column=1, value="Billing rate")
    ws.cell(row=9, column=2, value=rate_str)
    ws.cell(row=10, column=1, value="Billing fee (USD)")
    ws.cell(row=10, column=2, value=billing_fee)

    # ── Rows 12–15: rate change section (attribute D) ──────────────
    ws.cell(row=12, column=1, value="Rate change (Attribute D)")
    prior = prior_quarter(quarter)
    prior_rate_str = "N/A — first period" if prior is None else billing_rate(prior)
    ws.cell(row=13, column=1, value="Prior period rate")
    ws.cell(row=13, column=2, value=prior_rate_str)
    ws.cell(row=14, column=1, value="Current period rate")
    ws.cell(row=14, column=2, value=rate_str)

    change_rng = seeded_rng(spec, f"dc9_{quarter}_rate_change_date")
    if rate_changed_this_quarter(quarter):
        change_date_str = date_in_quarter(change_rng, quarter).isoformat()
    else:
        change_date_str = "N/A — no rate change"
    ws.cell(row=15, column=1, value="Change effective date")
    ws.cell(row=15, column=2, value=change_date_str)

    if not rate_changed_this_quarter(quarter):
        amendment_text = "N/A — no rate change"
    elif qc.defect == "dc9_rate_change_without_amendment":
        amendment_text = "NO AMENDMENT FILED"
    else:
        amendment_text = f"Amendment on file — ref DOC-{spec.year}-{quarter}"
    ws.cell(row=16, column=1, value="Supporting amendment")
    ws.cell(row=16, column=2, value=amendment_text)

    # ── Rows 18–22: asset roll-forward (attribute E) ────────────────
    ws.cell(row=18, column=1, value="Asset roll-forward (Attribute E)")
    rf_rng = seeded_rng(spec, f"dc9_{quarter}_roll_forward")
    opening = rf_rng.randint(int(asset_value * 0.85), int(asset_value * 0.95))
    additions = rf_rng.randint(int(asset_value * 0.03), int(asset_value * 0.08))
    retirements = (
        opening + additions - asset_value
    )  # closing tie: opening+additions-retirements=asset_value
    ws.cell(row=19, column=1, value="Opening balance (USD)")
    ws.cell(row=19, column=2, value=opening)
    ws.cell(row=20, column=1, value="Additions (USD)")
    ws.cell(row=20, column=2, value=additions)
    ws.cell(row=21, column=1, value="Retirements (USD)")
    ws.cell(row=21, column=2, value=retirements)
    ws.cell(row=22, column=1, value="Closing balance (USD)")
    ws.cell(row=22, column=2, value=asset_value)  # ties by construction

    # ── Rows 24–27: ownership share table (attribute F) ─────────────
    ws.cell(row=24, column=1, value="Ownership share (Attribute F)")
    ws.cell(row=25, column=1, value="Limited partner")
    ws.cell(row=25, column=2, value="Committed %")
    ws.cell(row=25, column=3, value="Effective %")
    for i, (lp_name, pct) in enumerate(_OWNERSHIP_LPS, start=26):
        ws.cell(row=i, column=1, value=lp_name)
        ws.cell(row=i, column=2, value=pct)
        ws.cell(row=i, column=3, value=pct)

    return wb
