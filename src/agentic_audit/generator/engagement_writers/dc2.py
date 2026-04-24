"""DC-2 quarterly workpaper — all 4 attributes in one worksheet.

One file per quarter. Each file carries every piece of evidence needed
to approve or deny every DC-2 attribute for that quarter.

Section layout (single worksheet, "DC-2 Variance"):

* Rows 1–3  — Header (entity, quarter, threshold %)
* Rows 5–7  — Attribute A : upstream feed total vs workbook total + tie flag
* Rows 9–15 — Attribute B + C : variance table, one row per revenue stream
                                 (prior / current / variance / % / above-thr? /
                                  explanation / source tie)
* Row 17    — Attribute D : reviewer initials + date

Defect handling:

* ``dc2_variance_boundary``               → upstream feed total diverges from
                                             workbook total; attribute A fails.
* ``dc2_variance_no_explanation``          → at least one above-threshold row
                                             has a blank explanation;
                                             attribute B fails.
* ``dc2_variance_explanation_inadequate``  → above-threshold rows have text,
                                             but ``source tie`` = ``No``;
                                             attribute C fails.
* ``none``                                 → every row ties; all attributes pass.

Attribute D is always clean (reviewer signs every quarter) in the §5 plan.
"""

from __future__ import annotations

from openpyxl import Workbook

from agentic_audit.generator.engagement_writers.common import (
    date_in_quarter,
    pick_initials,
    quarter_end_date,
    seeded_rng,
)
from agentic_audit.models.engagement import EngagementSpec, Quarter, quarter_control

# Revenue streams — fixed list, emitted in this order every quarter.
_REVENUE_STREAMS: tuple[str, ...] = (
    "Management Fees",
    "Performance Fees",
    "Interest Income",
    "Dividend Income",
    "Other Income",
)

# Threshold for flagging a variance — absolute % of prior period value.
_VARIANCE_THRESHOLD_PCT = 5.0


def render_dc2_quarter(spec: EngagementSpec, quarter: Quarter) -> Workbook:
    """Emit the DC-2 W/P for one quarter. Byte-deterministic per spec.seed.

    Returns a Workbook with one sheet named ``"DC-2 Variance"``.
    """
    qc = quarter_control(spec, "DC-2", quarter)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "DC-2 Variance"

    # ── Rows 1–3: header ────────────────────────────────────────────
    ws.cell(row=1, column=1, value="Entity")
    ws.cell(row=1, column=2, value=spec.entity_name)
    ws.cell(row=2, column=1, value="Quarter")
    ws.cell(row=2, column=2, value=f"{quarter} {spec.year}")
    ws.cell(row=3, column=1, value="Variance threshold (%)")
    ws.cell(row=3, column=2, value=_VARIANCE_THRESHOLD_PCT)

    # ── Compute per-stream variance data (seeded, deterministic) ─────
    stream_rng = seeded_rng(spec, f"dc2_{quarter}_streams")
    rows_data: list[tuple[str, int, int, int, float, bool]] = []
    for name in _REVENUE_STREAMS:
        prior = stream_rng.randint(500_000, 5_000_000)
        # Most streams move ±3 %; one or two exceed the 5 % threshold
        # so attribute B/C always have material content to examine.
        drift_pct = stream_rng.uniform(-0.12, 0.12)
        current = int(prior * (1 + drift_pct))
        variance = current - prior
        pct_variance = (variance / prior) * 100.0 if prior else 0.0
        above_threshold = abs(pct_variance) > _VARIANCE_THRESHOLD_PCT
        rows_data.append((name, prior, current, variance, pct_variance, above_threshold))

    workbook_total = sum(current for _, _, current, _, _, _ in rows_data)

    # ── Rows 5–7: attribute A (data load tie-out) ────────────────────
    ws.cell(row=5, column=1, value="Current-period data load (Attribute A)")
    # Upstream feed total agrees with workbook total unless attr A defect
    if qc.defect == "dc2_variance_boundary":
        upstream_total = workbook_total + stream_rng.randint(1_000, 10_000)
        tie_flag = "No"
    else:
        upstream_total = workbook_total
        tie_flag = "Yes"
    ws.cell(row=6, column=1, value="Upstream feed total (USD)")
    ws.cell(row=6, column=2, value=upstream_total)
    ws.cell(row=7, column=1, value="Workbook total (USD)")
    ws.cell(row=7, column=2, value=workbook_total)
    ws.cell(row=7, column=3, value=f"Ties to feed: {tie_flag}")

    # ── Row 9: variance-table header ─────────────────────────────────
    ws.cell(row=9, column=1, value="Revenue stream")
    ws.cell(row=9, column=2, value="Prior (USD)")
    ws.cell(row=9, column=3, value="Current (USD)")
    ws.cell(row=9, column=4, value="Variance (USD)")
    ws.cell(row=9, column=5, value="% variance")
    ws.cell(row=9, column=6, value="Above threshold?")
    ws.cell(row=9, column=7, value="Explanation")
    ws.cell(row=9, column=8, value="Source tie?")

    # ── Rows 10–14: one row per revenue stream ──────────────────────
    # Explanation + source-tie per defect type:
    #   none → always populated, source tie = Yes
    #   dc2_variance_no_explanation → first above-threshold row has blank
    #                                 explanation; others populated
    #   dc2_variance_explanation_inadequate → above-threshold rows have
    #                                          text but source tie = No
    no_expl_applied = False
    for i, (name, prior, current, variance, pct, above) in enumerate(rows_data, start=10):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=prior)
        ws.cell(row=i, column=3, value=current)
        ws.cell(row=i, column=4, value=variance)
        ws.cell(row=i, column=5, value=round(pct, 2))
        ws.cell(row=i, column=6, value="Yes" if above else "No")

        if not above:
            ws.cell(row=i, column=7, value="N/A — below threshold")
            ws.cell(row=i, column=8, value="N/A")
            continue

        # Above-threshold row — explanation + source-tie depend on defect
        if qc.defect == "dc2_variance_no_explanation" and not no_expl_applied:
            ws.cell(row=i, column=7, value="")  # missing — the defect
            ws.cell(row=i, column=8, value="No — no explanation provided")
            no_expl_applied = True
        elif qc.defect == "dc2_variance_explanation_inadequate":
            ws.cell(
                row=i,
                column=7,
                value=f"{name} movement driven by Q{quarter[1]} market fluctuation",
            )
            ws.cell(row=i, column=8, value="No")  # doesn't tie to upstream source
        else:
            ws.cell(
                row=i,
                column=7,
                value=(
                    f"{name} movement reconciled to GL feed — "
                    f"{'inflow' if variance > 0 else 'outflow'} matches subledger"
                ),
            )
            ws.cell(row=i, column=8, value="Yes")

    # ── Row 17: attribute D (reviewer sign-off) ─────────────────────
    rev_rng = seeded_rng(spec, f"dc2_{quarter}_reviewer")
    rev_initials = pick_initials(rev_rng)
    rev_date = date_in_quarter(rev_rng, quarter)
    ws.cell(row=17, column=1, value="Reviewer (Attribute D)")
    ws.cell(row=17, column=2, value=f"{rev_initials} — {rev_date.isoformat()}")

    # Keep the as-of date available (below, row 19) for agent tooling parity
    # with DC-9's header format.
    ws.cell(row=19, column=1, value="As-of date")
    ws.cell(row=19, column=2, value=quarter_end_date(quarter).isoformat())

    return wb
