"""Workbook structure writer.

Consumes a ScenarioSpec and produces a single-DC-sheet openpyxl.Workbook
matching the 9-section layout documented in
privateDocs/reference_templates/notes_on_structure.md §2.

This module emits structure + labels + placeholder markers only.
Task 3 replaces the ``<placeholder>`` cells with seeded fake data.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from agentic_audit.models.scenario import PatternType, ScenarioSpec

# Pattern-type dispatch tables. Extending to a new pattern means one
# entry here plus matching Literal additions in models.scenario.
_ATTRIBUTE_COUNT: dict[PatternType, int] = {
    "signoff_with_tieout": 6,
    "variance_detection": 4,
}

_SAMPLE_SIZE_LABEL: dict[PatternType, str] = {
    "signoff_with_tieout": "2 Quarters",
    "variance_detection": "2 months",
}

_FREQUENCY_LABEL: dict[PatternType, str] = {
    "signoff_with_tieout": "Quarterly",
    "variance_detection": "Monthly",
}

_SAMPLE_COUNT: dict[PatternType, int] = {
    "signoff_with_tieout": 2,
    "variance_detection": 2,
}


@dataclass
class SheetCursor:
    """Tracks the current row while emitting sections to a worksheet.

    Section emitters call ``advance(n)`` after writing, so the next
    section starts wherever the previous one ended — no hard-coded rows.
    """

    sheet: Worksheet
    row: int = 1

    def write(self, column: int, value: Any) -> None:
        """Write a value at (self.row, column)."""
        self.sheet.cell(row=self.row, column=column, value=value)

    def merge_row(self, start_col: int, end_col: int) -> None:
        """Merge cells on the current row from start_col to end_col inclusive."""
        self.sheet.merge_cells(
            start_row=self.row,
            start_column=start_col,
            end_row=self.row,
            end_column=end_col,
        )

    def advance(self, n: int = 1) -> None:
        """Move the cursor forward by n rows."""
        self.row += n


def _emit_header(cursor: SheetCursor, spec: ScenarioSpec) -> None:
    """Rows 1-7: fixed-position header block with label/value pairs."""
    # Row 1
    cursor.write(1, "Workpaper ref.:")
    cursor.merge_row(1, 2)
    cursor.write(3, "<workpaper_ref>")
    cursor.write(5, "Tests of the control:")
    cursor.merge_row(5, 6)
    cursor.write(7, f"{spec.control_id}: <control_name>")
    cursor.merge_row(7, 14)
    cursor.advance(1)
    # Row 2
    cursor.write(1, "Entity name:")
    cursor.merge_row(1, 2)
    cursor.write(3, "<entity_name>")
    cursor.write(5, "Related SCOT:")
    cursor.merge_row(5, 6)
    cursor.write(7, "<scot_name>")
    cursor.merge_row(7, 14)
    cursor.advance(1)
    # Row 3
    cursor.write(1, "Applicable GAAS:")
    cursor.merge_row(1, 2)
    cursor.write(3, "<gaas>")
    cursor.write(5, "Nature of the engagement:")
    cursor.merge_row(5, 6)
    cursor.write(7, "<engagement_nature>")
    cursor.merge_row(7, 14)
    cursor.advance(1)
    # Row 4
    cursor.write(1, "Year-end date:")
    cursor.merge_row(1, 2)
    cursor.write(3, "<year_end_date>")
    cursor.write(5, "Operating effectiveness conclusion:")
    cursor.merge_row(5, 6)
    cursor.write(7, "<effectiveness_conclusion>")
    cursor.merge_row(7, 10)
    cursor.advance(1)
    # Row 5
    cursor.write(1, "Prepared by")
    cursor.write(3, "<preparer_initials>")
    cursor.advance(1)
    # Row 6
    cursor.write(1, "Reviewed by ")
    cursor.write(3, "<reviewer_1_initials>")
    cursor.advance(1)
    # Row 7
    cursor.write(1, "Reviewed by ")
    cursor.merge_row(1, 2)
    cursor.write(3, "<reviewer_2_initials>")
    cursor.advance(1)


def _emit_control_info(cursor: SheetCursor, spec: ScenarioSpec) -> None:
    """Rows 8-15: control-info block with banner + field rows."""
    # Banner
    cursor.write(1, "Information about the control")
    cursor.merge_row(1, 14)
    cursor.advance(2)  # row 9 intentionally blank
    # Row 10: Workpaper reference of SCOT form
    cursor.write(2, "Workpaper reference of SCOT form")
    cursor.merge_row(2, 6)
    cursor.write(7, "<scot_form_ref>")
    cursor.merge_row(7, 8)
    cursor.advance(1)
    # Row 11: Control description
    cursor.write(2, "Control description")
    cursor.merge_row(2, 3)
    cursor.write(4, "<control_description>")
    cursor.merge_row(4, 14)
    cursor.advance(2)  # row 12 blank
    # Row 13: Control type + Frequency
    cursor.write(2, "Control type")
    cursor.merge_row(2, 3)
    cursor.write(4, "<control_type>")
    cursor.merge_row(4, 6)
    cursor.write(10, "Frequency")
    cursor.write(11, _FREQUENCY_LABEL[spec.pattern_type])
    cursor.advance(1)
    # Row 14: Review or monitoring + Related IT application(s)
    cursor.write(2, "Review or monitoring control?")
    cursor.merge_row(2, 3)
    cursor.write(4, "<review_or_monitoring>")
    cursor.merge_row(4, 6)
    cursor.write(10, "Related IT application(s)")
    cursor.write(11, "<it_applications>")
    cursor.advance(2)  # skip row 15 into row 16


def _emit_ipe(cursor: SheetCursor, spec: ScenarioSpec) -> None:
    """Rows 16-20: IPE block."""
    # Banner
    cursor.write(1, "Information produced by the entity (IPE) used in the execution of the control")
    cursor.merge_row(1, 14)
    cursor.advance(2)  # row 17 blank
    # Row 18
    cursor.write(
        2,
        "Is any information produced by the entity used in the execution of this control?",
    )
    cursor.write(7, "<ipe_yes_no>")
    cursor.advance(1)
    # Row 19
    cursor.write(
        2,
        "Reference the workpaper where we document how the IPE risks are addressed",
    )
    cursor.write(7, "<ipe_risk_ref>")
    cursor.advance(2)  # row 20 blank, advance into row 21


def _emit_multi_instance(cursor: SheetCursor, spec: ScenarioSpec) -> None:
    """Rows 21-29: multi-instance / configurability block."""
    # Banner
    cursor.write(
        1,
        "Determine that the multiple instances of the same IT application have "
        "the same functionality as the test sample",
    )
    cursor.merge_row(1, 14)
    cursor.advance(2)  # row 22 blank
    # Row 23
    cursor.write(
        2,
        "Is this test performed on a single instance of a multiple-instance IT application "
        "and intended to serve as the testing for all instances?",
    )
    cursor.advance(2)  # row 24 blank
    # Row 25
    cursor.write(2, "Is this control configurable?")
    cursor.write(6, "<configurable_yes_no>")
    cursor.advance(2)  # row 26 blank
    # Row 27
    cursor.write(
        2,
        "Reference the workpaper or describe procedures performed to validate "
        "the IT application's report / configuration settings",
    )
    cursor.write(6, "<report_validation>")
    cursor.advance(2)  # row 28 blank
    # Row 29
    cursor.write(
        2,
        "Describe the procedures performed to evaluate the appropriateness of "
        "the scope and nature of the report",
    )
    cursor.advance(1)


def _emit_sample_size(cursor: SheetCursor, spec: ScenarioSpec) -> None:
    """Rows 30-32: sample size block."""
    cursor.advance(1)  # row 30 blank
    # Row 31
    cursor.write(2, "Sample size?")
    cursor.write(5, _SAMPLE_SIZE_LABEL[spec.pattern_type])
    cursor.advance(1)
    # Row 32
    cursor.write(
        2,
        "Description of the procedures performed to determine the population for testing…",
    )
    cursor.write(5, "<population_description>")
    cursor.advance(1)


def _emit_testing_procedures(cursor: SheetCursor, spec: ScenarioSpec) -> None:
    """Testing procedures banner + column header + N attribute rows."""
    # Banner
    cursor.write(1, "Testing procedures")
    cursor.merge_row(1, 14)
    cursor.advance(2)  # skip blank
    # Column headers
    cursor.write(3, "Design attributes")
    cursor.write(7, "ToC procedures to be performed")
    cursor.advance(1)
    # Attribute rows — count depends on pattern
    attribute_count = _ATTRIBUTE_COUNT[spec.pattern_type]
    attribute_letters = "ABCDEF"[:attribute_count]
    for letter in attribute_letters:
        cursor.write(1, "Del")
        cursor.write(2, letter)
        cursor.write(3, f"<attribute_{letter}_description>")
        cursor.write(7, f"<attribute_{letter}_toc_procedure>")
        cursor.advance(1)
    cursor.advance(1)  # trailing blank


def _emit_sample_grid(cursor: SheetCursor, spec: ScenarioSpec) -> None:
    """Result of procedures: banner + header row + sample rows."""
    # Banner
    cursor.write(1, "Result of procedures")
    cursor.merge_row(1, 14)
    cursor.advance(2)  # skip blank
    # Header row
    cursor.write(2, "Sample item #")
    cursor.write(3, "Description of sample item")
    cursor.write(6, "Date/period")
    cursor.write(7, "W/P references (if needed)")
    # Attribute column headers H–M
    attribute_count = _ATTRIBUTE_COUNT[spec.pattern_type]
    attribute_letters = "ABCDEF"[:attribute_count]
    for idx, letter in enumerate(attribute_letters):
        cursor.write(8 + idx, letter)  # H=8, I=9, J=10, K=11, L=12, M=13
    cursor.advance(1)
    # Sample rows
    sample_count = _SAMPLE_COUNT[spec.pattern_type]
    for i in range(1, sample_count + 1):
        cursor.write(1, "Del")
        cursor.write(2, i)
        cursor.write(3, f"<sample_{i}_description>")
        cursor.write(6, f"<sample_{i}_period>")
        cursor.write(7, f"<sample_{i}_wp_ref>")
        for idx in range(attribute_count):
            cursor.write(8 + idx, "<tickmark>")
        cursor.advance(1)
    cursor.advance(1)  # trailing blank


def _emit_tickmark_key(cursor: SheetCursor, spec: ScenarioSpec) -> None:
    """Tickmark key legend rows."""
    cursor.write(2, "Tickmark key")
    cursor.advance(1)
    cursor.write(2, "a")
    cursor.write(3, "Attribute satisfied without exception")
    cursor.advance(1)
    cursor.write(2, "X")
    cursor.write(3, "Attribute satisfied with exception")
    cursor.advance(2)  # trailing blank


def _emit_conclusion(cursor: SheetCursor, spec: ScenarioSpec) -> None:
    """Conclusion banner + 7 field rows."""
    # Banner
    cursor.write(1, "Conclusion")
    cursor.merge_row(1, 14)
    cursor.advance(2)  # skip blank
    # 7 field rows (label in col B, value in col H)
    fields: tuple[tuple[str, str], ...] = (
        ("Were exceptions noted during testing?", "<exceptions_noted>"),
        (
            "Do we have persuasive evidence that the exceptions are random?",
            "<exceptions_random>",
        ),
        (
            "Describe our understanding of the nature and cause of the exceptions",
            "<exception_nature>",
        ),
        (
            "Document how we extended our sample to address the random control exception",
            "<sample_extension>",
        ),
        ("Did we identify and test effective compensating controls?", "<compensating_yes_no>"),
        (
            "Reference the name(s) of the effective compensating control(s)",
            "<compensating_names>",
        ),
        (
            "Describe if exceptions are determined to be deficiencies and refer to "
            "the deficiency evaluation workpaper",
            "<deficiency_refs>",
        ),
        (
            "Operating effectiveness conclusion for the period of reliance",
            "<effectiveness_conclusion>",
        ),
    )
    for label, placeholder in fields:
        cursor.write(2, label)
        cursor.write(8, placeholder)
        cursor.advance(1)


def render_toc_sheet(spec: ScenarioSpec) -> Workbook:
    """Build a single-sheet TOC workbook from a ScenarioSpec.

    Returns an in-memory openpyxl Workbook. Callers are responsible for
    persistence (``wb.save(path)``). Task 3 will replace the placeholder
    markers with seeded fake data.
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None, "New Workbook always has an active sheet"
    ws.title = spec.control_id
    assert len(ws.title) <= 31, f"Sheet name {ws.title!r} exceeds Excel's 31-char limit"

    cursor = SheetCursor(sheet=ws, row=1)
    _emit_header(cursor, spec)
    _emit_control_info(cursor, spec)
    _emit_ipe(cursor, spec)
    _emit_multi_instance(cursor, spec)
    _emit_sample_size(cursor, spec)
    _emit_testing_procedures(cursor, spec)
    _emit_sample_grid(cursor, spec)
    _emit_tickmark_key(cursor, spec)
    _emit_conclusion(cursor, spec)

    return wb
