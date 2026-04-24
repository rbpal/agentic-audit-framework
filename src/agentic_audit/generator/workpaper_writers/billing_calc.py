"""Billing-calculation supporting workpaper — DC-9 scenarios.

A minimal 2-column workbook (``label`` | ``value``) documenting the
billing computation the TOC claims to have tied out. Emits fields a
real auditor would expect to see in a supporting schedule:

* Entity name
* Period (quarter + year)
* Asset value
* Billing rate
* Billing fee = assets × rate
* Preparer initials + date

Task 12 scope is the writer itself — produces deterministic content
from ``ScenarioSpec.seed`` with no cross-reference to the TOC's
narrative. Task 13 will introduce cross-file consistency rules (e.g.,
``figure_mismatch`` makes the billing fee *disagree* with the TOC).
"""

from __future__ import annotations

import random

from openpyxl import Workbook

from agentic_audit.generator import fake_data
from agentic_audit.models.scenario import ScenarioSpec, WorkpaperSpec

# Discrete rate pool — mirrors what a real rate card offers. Rate choice
# is deterministic per seed; all downstream values (fee) compute from it.
_BILLING_RATES: tuple[str, ...] = ("0.25%", "0.50%", "0.75%", "1.00%")


def render_billing_calc(spec: ScenarioSpec, wp_spec: WorkpaperSpec) -> Workbook:
    """Produce a billing-calculation ``.xlsx`` workbook.

    Args:
        spec: the parent scenario — drives entity, period, asset value, rate
        wp_spec: the workpaper specification — header metadata only today;
                 extensible in Task 13 for cross-file consistency rules

    Returns:
        An openpyxl Workbook with a single sheet named ``Billing Calc``.
        Byte-deterministic given the same ``(spec, wp_spec)`` pair.
    """
    rng = random.Random(spec.seed)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Billing Calc"

    entity = fake_data.fake_entity_name(rng)
    year_end = fake_data.fake_year_end_date(rng, spec.quarter)
    period_str = f"{spec.quarter} {year_end.year}"
    asset_value = fake_data.fake_asset_value_usd(rng)
    rate_str = rng.choice(_BILLING_RATES)
    rate_decimal = float(rate_str.rstrip("%")) / 100.0
    billing_fee = int(asset_value * rate_decimal)
    preparer_initials = fake_data.fake_initials(rng)
    prep_date = fake_data.fake_year_end_date(rng, spec.quarter)

    rows: list[tuple[str, object | None]] = [
        (f"Billing Calculation — {wp_spec.toc_reference_code}", None),
        ("Entity", entity),
        ("Period", period_str),
        ("Asset value (USD)", asset_value),
        ("Billing rate", rate_str),
        ("Billing fee (USD)", billing_fee),
        ("Prepared by", preparer_initials),
        ("Date prepared", prep_date.isoformat()),
    ]
    for r, (label, value) in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=label)
        if value is not None:
            ws.cell(row=r, column=2, value=value)

    return wb
