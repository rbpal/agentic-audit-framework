"""Step-1 corpus → bronze ingest helpers (smoke-test scope).

Two extractors:

* :func:`extract_workpaper_rows` — yields one :class:`WorkpaperRow` per
  non-empty xlsx row. Produces records that map 1:1 to
  ``audit_dev.bronze.workpapers_raw`` columns.
* :func:`extract_toc_record` — yields a single :class:`TocRecord` per
  TOC json file. Maps 1:1 to ``audit_dev.bronze.tocs_raw`` columns.

Both extractors compute :func:`file_sha256` over the on-disk bytes — that's
the ``file_hash`` lineage anchor stored in bronze. Distinct from
:func:`agentic_audit.generator.content_hash.content_hash`, which canonicalises
xlsx zip metadata. Bronze stores the *file-as-delivered* hash; auditors verify
against that, not the content hash.
"""

from __future__ import annotations

import hashlib
import json
import re
from collections.abc import Iterator, Sequence
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl

# Filenames in eval/gold_scenarios/ follow ``dc{n}_Q{q}[_ref].(xlsx|json)``.
_FILENAME_RE = re.compile(r"^dc(?P<n>\d+)_Q(?P<q>[1-4])(?:_ref)?\.(?:xlsx|json)$", re.IGNORECASE)


@dataclass(frozen=True)
class WorkpaperRow:
    """One row destined for ``audit_dev.bronze.workpapers_raw``."""

    source_path: str
    file_hash: str
    engagement_id: str
    sheet_name: str
    row_index: int
    raw_data: dict[str, str]
    ingested_at: datetime
    ingested_by: str


@dataclass(frozen=True)
class TocRecord:
    """One record destined for ``audit_dev.bronze.tocs_raw``."""

    source_path: str
    file_hash: str
    engagement_id: str
    control_id: str
    quarter: str
    raw_json: str
    ingested_at: datetime
    ingested_by: str


def file_sha256(path: Path) -> str:
    """SHA-256 hex digest of the file's on-disk bytes."""
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def parse_corpus_filename(path: Path) -> tuple[str, str]:
    """Extract ``(control_id, quarter)`` from a Step-1 corpus filename.

    Raises :class:`ValueError` for filenames that don't match the convention —
    rather than silently inferring, because bronze lineage demands certainty.
    """
    m = _FILENAME_RE.match(path.name)
    if not m:
        raise ValueError(
            f"Filename {path.name!r} does not match ``dc<n>_Q<1-4>[_ref].(xlsx|json)``."
        )
    return f"DC-{int(m['n'])}", f"Q{m['q']}"


def _row_to_raw_data(row: Sequence[object]) -> dict[str, str]:
    """Stringify a sheet row into a column-index → value map.

    Uses zero-padded ``col_NN`` keys instead of header names because
    Step-1 workpapers are mixed-layout (label/value pairs interleaved
    with section banners) — there is no consistent header row to anchor
    the keys to. Empty cells are omitted to keep the map dense.
    """
    out: dict[str, str] = {}
    for i, value in enumerate(row):
        if value is None:
            continue
        out[f"col_{i:02d}"] = str(value)
    return out


def extract_workpaper_rows(
    path: Path,
    *,
    engagement_id: str,
    ingested_by: str,
    ingested_at: datetime,
) -> Iterator[WorkpaperRow]:
    """Yield one :class:`WorkpaperRow` per non-empty xlsx row.

    ``row_index`` is 1-based. Empty rows (every cell ``None``) are
    skipped — they're a presentational quirk in the source xlsx and
    carry no evidence.
    """
    file_hash = file_sha256(path)
    source_path = str(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                raw_data = _row_to_raw_data(row)
                if not raw_data:
                    continue
                yield WorkpaperRow(
                    source_path=source_path,
                    file_hash=file_hash,
                    engagement_id=engagement_id,
                    sheet_name=sheet_name,
                    row_index=row_idx,
                    raw_data=raw_data,
                    ingested_at=ingested_at,
                    ingested_by=ingested_by,
                )
    finally:
        wb.close()


def extract_toc_record(
    path: Path,
    *,
    engagement_id: str,
    ingested_by: str,
    ingested_at: datetime,
) -> TocRecord:
    """Read one TOC json file as a :class:`TocRecord`.

    Validates that the json is parseable (so we fail fast on a corrupt
    file at ingest time rather than at silver-tier ETL) but stores the
    verbatim text in ``raw_json`` — the bronze contract is byte-faithful
    capture, not normalised payload.
    """
    raw_json = path.read_text(encoding="utf-8")
    json.loads(raw_json)  # parse-only; result discarded — we keep verbatim text
    control_id, quarter = parse_corpus_filename(path)
    return TocRecord(
        source_path=str(path),
        file_hash=file_sha256(path),
        engagement_id=engagement_id,
        control_id=control_id,
        quarter=quarter,
        raw_json=raw_json,
        ingested_at=ingested_at,
        ingested_by=ingested_by,
    )


def discover_corpus(root: Path) -> tuple[list[Path], list[Path]]:
    """Return ``(workpaper_xlsx_paths, toc_json_paths)`` under ``root``.

    Expects the canonical layout::

        <root>/workpapers/dc{n}_Q{q}_ref.xlsx
        <root>/tocs/dc{n}_Q{q}.json

    Files not matching the regex (e.g. the human-readable
    ``engagement_toc_ref.xlsx``) are excluded from the smoke ingest —
    they're reference artefacts, not corpus inputs.
    """
    workpapers = sorted(
        p for p in (root / "workpapers").glob("*.xlsx") if _FILENAME_RE.match(p.name)
    )
    tocs = sorted(p for p in (root / "tocs").glob("*.json") if _FILENAME_RE.match(p.name))
    return workpapers, tocs
